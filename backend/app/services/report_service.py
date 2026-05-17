"""Hisobotlarni generatsiya qilish — PDF va Excel."""
import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def build_excel_report(title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    """Excel hisobotni byte-stringda qaytaradi."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    # Title
    ws.cell(1, 1, title).font = Font(bold=True, size=16)
    ws.cell(2, 1, f"Generated: {datetime.now():%Y-%m-%d %H:%M}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Headers
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(4, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(r_idx, c_idx, val)
            c.border = border
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_report(title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    """PDF hisobotni byte-stringda qaytaradi."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=30,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Paragraph(f"Yaratilgan: {datetime.now():%Y-%m-%d %H:%M}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [headers] + [[str(v) if v is not None else "" for v in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1677FF")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()
